"""Render report (title, columns, rows) to CSV / Excel / PDF HttpResponses."""
import csv
import io

from django.http import HttpResponse


def to_csv(title, columns, rows):
    resp = HttpResponse(content_type="text/csv")
    resp["Content-Disposition"] = f'attachment; filename="{_slug(title)}.csv"'
    writer = csv.writer(resp)
    writer.writerow(columns)
    for row in rows:
        writer.writerow(row)
    return resp


def to_excel(title, columns, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    header_fill = PatternFill("solid", fgColor="4F46E5")
    bold_white = Font(bold=True, color="FFFFFF")
    ws.append(columns)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = bold_white
    for row in rows:
        ws.append(list(row))
    for i, col in enumerate(columns, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(
            14, len(str(col)) + 4
        )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{_slug(title)}.xlsx"'
    return resp


def to_pdf(title, columns, rows):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles["Title"]), Spacer(1, 12)]
    data = [columns] + [[str(c) for c in row] for row in rows]
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{_slug(title)}.pdf"'
    return resp


def _slug(title):
    return title.lower().replace(" ", "_")


EXPORTERS = {"csv": to_csv, "excel": to_excel, "xlsx": to_excel, "pdf": to_pdf}


def export(fmt, title, columns, rows):
    fn = EXPORTERS.get(fmt)
    if not fn:
        raise ValueError(f"Unknown export format: {fmt}")
    return fn(title, columns, rows)
